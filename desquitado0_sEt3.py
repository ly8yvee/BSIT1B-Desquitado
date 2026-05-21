
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

filename = "ordersDB.xlsx"

window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")

selected_row = None

def display_data():
    table.delete(*table.get_children())

    wb = op.load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)

    wb.close()


def select_record(event):
    global selected_row

    selected_row = table.focus()

    if selected_row:
        values = table.item(selected_row, "values")

        cname_entry.delete(0, tk.END)
        product_entry.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        cname_entry.insert(0, values[1])
        product_entry.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])


def update_record():
    global selected_row

    if not selected_row:
        messagebox.showerror("Error", "Select a record first")
        return

    cname = cname_entry.get()
    product = product_entry.get()
    qty = qty_entry.get()
    price = price_entry.get()

    if cname == "" or product == "" or qty == "" or price == "":
        messagebox.showerror("Error", "Complete all fields")
        return

    try:
        qty = int(qty)
        price = float(price)
    except:
        messagebox.showerror("Error", "Quantity and Price must be numeric")
        return

    total = qty * price

    values = table.item(selected_row, "values")
    order_id = values[0]

    wb = op.load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):

        if row[0].value == int(order_id):

            row[1].value = cname
            row[2].value = product
            row[3].value = qty
            row[4].value = price
            row[5].value = total

    wb.save(filename)
    wb.close()

    display_data()

    messagebox.showinfo("Success", "Record Updated")


def delete_record():
    global selected_row

    if not selected_row:
        messagebox.showerror("Error", "Select a record first")
        return

    values = table.item(selected_row, "values")
    order_id = values[0]

    wb = op.load_workbook(filename)
    ws = wb.active

    for row in range(2, ws.max_row + 1):

        if ws.cell(row=row, column=1).value == int(order_id):
            ws.delete_rows(row)
            break

    wb.save(filename)
    wb.close()

    display_data()

    messagebox.showinfo("Success", "Record Deleted")


# =========================
# GUI
# =========================

title = tk.Label(window, text="Simple Ordering System",
                 font=("Times New Roman", 14, "bold"),
                 bg="lightblue")

title.grid(row=0, column=0, columnspan=6)

genframe = tk.Frame(window, bg="lightblue",
                    bd=2, relief="groove")

genframe.grid(row=1, column=0,
              columnspan=7,
              padx=10,
              pady=10)

# Customer Name
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1,
                 columnspan=2,
                 padx=10,
                 pady=(10, 0))

cname_label = tk.Label(genframe,
                       text="Customer Name",
                       font=("Poppins", 10, "italic"),
                       bg="lightblue")

cname_label.grid(row=3, column=1, columnspan=2)

# Product
product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3,
                   columnspan=2,
                   padx=10,
                   pady=(10, 0))

product_label = tk.Label(genframe,
                         text="Product",
                         font=("Poppins", 10, "italic"),
                         bg="lightblue")

product_label.grid(row=3, column=3, columnspan=2)

# Quantity
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1,
               columnspan=2,
               padx=10,
               pady=(10, 0))

qty_label = tk.Label(genframe,
                     text="Quantity",
                     font=("Poppins", 10, "italic"),
                     bg="lightblue")

qty_label.grid(row=5, column=1, columnspan=2)

# Price
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3,
                 columnspan=2,
                 padx=10,
                 pady=(10, 0))

price_label = tk.Label(genframe,
                       text="Price",
                       font=("Poppins", 10, "italic"),
                       bg="lightblue")

price_label.grid(row=5, column=3, columnspan=2)

# Buttons
update_btn = tk.Button(window,
                       text="Update",
                       font=("Poppins", 12, "bold"),
                       bg="lightgreen",
                       command=update_record)

update_btn.grid(row=6, column=2)

delete_btn = tk.Button(window,
                       text="Delete",
                       bg="red",
                       fg="white",
                       font=("Poppins", 12, "bold"),
                       command=delete_record)

delete_btn.grid(row=6, column=3)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID",
             "Customer Name",
             "Product",
             "Quantity",
             "Price",
             "Total"),
    show="headings"
)

for headings in ("Order ID",
                 "Customer Name",
                 "Product",
                 "Quantity",
                 "Price",
                 "Total"):

    table.heading(headings, text=headings)

table.grid(row=7,
           column=0,
           columnspan=6,
           padx=10,
           pady=10)

table.bind("<ButtonRelease-1>", select_record)

# DISPLAY DATA
display_data()

window.mainloop()