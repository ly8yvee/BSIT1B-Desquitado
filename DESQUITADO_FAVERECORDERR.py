import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl as op
from datetime import datetime
import os

window = tk.Tk()
window.title("Profile Builder")
window.geometry("800x500")
window.configure(bg="lightgreen")

filename = "favorite_people.xlsx"

if os.path.exists(filename):
    workbook = op.load_workbook(filename)
    sheet = workbook.active
else:
    workbook = op.Workbook()
    sheet = workbook.active
    headers = ["ID", "First Name", "Middle Name", "Last Name", "Birth Year", "Age"]
    sheet.append(headers)
    workbook.save(filename)

title = tk.Label(window, text="Profile Builder", font=("Times New Roman", 14, "bold"), bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

genframe = tk.Frame(window, bg="lightgreen", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=6, padx=10, pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

tree = ttk.Treeview(window, columns=("ID", "First", "Middle", "Last", "Birth", "Age"), show="headings")

for col in ("ID", "First", "Middle", "Last", "Birth", "Age"):
    tree.heading(col, text=col)

tree.grid(row=6, column=0, columnspan=6)

selected = None
selected_row = None


def load_data():
    for item in tree.get_children():
        tree.delete(item)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert("", tk.END, values=row)


def clear_data():
    fname_entry.delete(0, tk.END)
    mname_entry.delete(0, tk.END)
    lname_entry.delete(0, tk.END)
    birth_entry.delete(0, tk.END)


def submit_data():
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if first == "" or last == "" or birth == "":
        messagebox.showerror("Error", "Fill all required fields")
        return

    try:
        birth = int(birth)
    except:
        messagebox.showerror("Error", "Birth year must be number")
        return

    age = datetime.now().year - birth
    new_id = sheet.max_row

    sheet.append([new_id, first, middle, last, birth, age])
    workbook.save(filename)

    load_data()
    clear_data()
    messagebox.showinfo("Success", "Saved")


def select_record(event):
    global selected, selected_row

    selected = tree.focus()
    values = tree.item(selected, "values")

    if not values:
        return

    selected_row = int(values[0]) + 1

    clear_data()

    fname_entry.insert(0, values[1])
    mname_entry.insert(0, values[2])
    lname_entry.insert(0, values[3])
    birth_entry.insert(0, values[4])


tree.bind("<<TreeviewSelect>>", select_record)


def update_data():
    if selected_row is None:
        messagebox.showerror("Error", "Select a record first")
        return

    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()

    try:
        birth = int(birth_entry.get())
    except:
        messagebox.showerror("Error", "Birth year must be number")
        return

    age = datetime.now().year - birth

    sheet.cell(row=selected_row, column=2).value = first
    sheet.cell(row=selected_row, column=3).value = middle
    sheet.cell(row=selected_row, column=4).value = last
    sheet.cell(row=selected_row, column=5).value = birth
    sheet.cell(row=selected_row, column=6).value = age

    workbook.save(filename)
    load_data()
    clear_data()
    messagebox.showinfo("Updated", "Record Updated")


def delete_data():
    if selected_row is None:
        messagebox.showerror("Error", "Select a record first")
        return

    sheet.delete_rows(selected_row)
    workbook.save(filename)

    load_data()
    clear_data()
    messagebox.showinfo("Deleted", "Record Deleted")

tk.Button(window, text="Submit", bg="pink", command=submit_data).grid(row=3, column=1)
tk.Button(window, text="Update", bg="orange", command=update_data).grid(row=3, column=2)
tk.Button(window, text="Delete", bg="red", fg="white", command=delete_data).grid(row=3, column=3)
tk.Button(window, text="Clear", bg="skyblue", command=clear_data).grid(row=3, column=4)

load_data()

window.mainloop()