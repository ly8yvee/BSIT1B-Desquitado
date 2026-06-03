import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
from openpyxl import Workbook
import os

window = tk.Tk()
window.title("Study Room Reservation System")
window.configure(bg="lightblue")

filename = "desquitado_database.xlsx"

if not os.path.exists(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservations"

    headings = [
        "Reservation ID",
        "Student Name",
        "Student ID",
        "Room Number",
        "Hours",
        "Rate Per Hour",
        "Total Cost"
    ]

    ws.append(headings)
    wb.save(filename)

# Generate Reservation ID
def generate_id():
    wb = op.load_workbook(filename)
    ws = wb.active

    row_count = ws.max_row
    reservation_id = f"C-{row_count:03d}"

    rid_entry.config(state="normal")
    rid_entry.delete(0, tk.END)
    rid_entry.insert(0, reservation_id)
    rid_entry.config(state="readonly")

# Compute Total Payment
def compute_total():
    try:
        hours = int(hours_entry.get())
        rate = float(rate_entry.get())

        total = hours * rate

        total_entry.config(state="normal")
        total_entry.delete(0, tk.END)
        total_entry.insert(0, str(total))
        total_entry.config(state="readonly")

    except:
        pass

def add_record():

    if (
        sname_entry.get() == "" or
        sid_entry.get() == "" or
        room_entry.get() == "" or
        hours_entry.get() == "" or
        rate_entry.get() == ""
    ):
        messagebox.showerror("Error", "Please fill all fields")
        return

    compute_total()

    wb = op.load_workbook(filename)
    ws = wb.active

    data = [
        rid_entry.get(),
        sname_entry.get(),
        sid_entry.get(),
        room_entry.get(),
        hours_entry.get(),
        rate_entry.get(),
        total_entry.get()
    ]

    ws.append(data)
    wb.save(filename)

    messagebox.showinfo("Success", "Reservation Added")

    display_records()
    clear_fields()
    generate_id()

def display_records():

    for row in table.get_children():
        table.delete(row)

    wb = op.load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)

def select_record(event):

    selected = table.focus()

    if not selected:
        return

    values = table.item(selected, "values")

    rid_entry.config(state="normal")
    rid_entry.delete(0, tk.END)
    rid_entry.insert(0, values[0])
    rid_entry.config(state="readonly")

    sname_entry.delete(0, tk.END)
    sname_entry.insert(0, values[1])

    sid_entry.delete(0, tk.END)
    sid_entry.insert(0, values[2])

    room_entry.delete(0, tk.END)
    room_entry.insert(0, values[3])

    hours_entry.delete(0, tk.END)
    hours_entry.insert(0, values[4])

    rate_entry.delete(0, tk.END)
    rate_entry.insert(0, values[5])

    total_entry.config(state="normal")
    total_entry.delete(0, tk.END)
    total_entry.insert(0, values[6])
    total_entry.config(state="readonly")

def update_record():

    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return

    compute_total()

    wb = op.load_workbook(filename)
    ws = wb.active

    selected_id = rid_entry.get()

    for row in ws.iter_rows(min_row=2):

        if row[0].value == selected_id:

            row[1].value = sname_entry.get()
            row[2].value = sid_entry.get()
            row[3].value = room_entry.get()
            row[4].value = hours_entry.get()
            row[5].value = rate_entry.get()
            row[6].value = total_entry.get()

    wb.save(filename)

    messagebox.showinfo("Success", "Record Updated")

    display_records()
    clear_fields()

def delete_record():

    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return

    confirm = messagebox.askyesno(
        "Confirm",
        "Do you want to delete this record?"
    )

    if confirm:

        wb = op.load_workbook(filename)
        ws = wb.active

        selected_id = rid_entry.get()

        for row in range(2, ws.max_row + 1):

            if ws.cell(row=row, column=1).value == selected_id:
                ws.delete_rows(row)
                break

        wb.save(filename)

        messagebox.showinfo("Success", "Record Deleted")

        display_records()
        clear_fields()

def clear_fields():

    sname_entry.delete(0, tk.END)
    sid_entry.delete(0, tk.END)
    room_entry.delete(0, tk.END)
    hours_entry.delete(0, tk.END)
    rate_entry.delete(0, tk.END)

    total_entry.config(state="normal")
    total_entry.delete(0, tk.END)
    total_entry.config(state="readonly")

title = tk.Label(
    window,
    text="STUDY ROOM RESERVATION SYSTEM",
    font=("Times New Roman", 16, "bold"),
    bg="lightblue"
)

title.grid(row=0, column=0, columnspan=6, pady=10)


genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=6, padx=10, pady=10)
rid_entry = tk.Entry(genframe, font=("Poppins", 12), state="readonly")
rid_entry.grid(row=0, column=1, padx=10, pady=5)

rid_label = tk.Label(
    genframe,
    text="Reservation ID",
    bg="lightblue"
)
rid_label.grid(row=0, column=0)

sname_entry = tk.Entry(genframe, font=("Poppins", 12))
sname_entry.grid(row=1, column=1, padx=10, pady=5)

sname_label = tk.Label(
    genframe,
    text="Student Name",
    bg="lightblue"
)
sname_label.grid(row=1, column=0)


sid_entry = tk.Entry(genframe, font=("Poppins", 12))
sid_entry.grid(row=2, column=1, padx=10, pady=5)

sid_label = tk.Label(
    genframe,
    text="Student ID",
    bg="lightblue"
)
sid_label.grid(row=2, column=0)

room_entry = tk.Entry(genframe, font=("Poppins", 12))
room_entry.grid(row=3, column=1, padx=10, pady=5)

room_label = tk.Label(
    genframe,
    text="Room Number",
    bg="lightblue"
)
room_label.grid(row=3, column=0)

hours_entry = tk.Entry(genframe, font=("Poppins", 12))
hours_entry.grid(row=4, column=1, padx=10, pady=5)

hours_label = tk.Label(
    genframe,
    text="Hours",
    bg="lightblue"
)
hours_label.grid(row=4, column=0)

rate_entry = tk.Entry(genframe, font=("Poppins", 12))
rate_entry.grid(row=5, column=1, padx=10, pady=5)

rate_label = tk.Label(
    genframe,
    text="Rate Per Hour",
    bg="lightblue"
)
rate_label.grid(row=5, column=0)

total_entry = tk.Entry(
    genframe,
    font=("Poppins", 12),
    state="readonly"
)

total_entry.grid(row=6, column=1, padx=10, pady=5)

total_label = tk.Label(
    genframe,
    text="Total Payment",
    bg="lightblue"
)
total_label.grid(row=6, column=0)

submit_btn = tk.Button(
    window,
    text="Submit",
    font=("Poppins", 12, "bold"),
    bg="lightpink",
    command=add_record
)

submit_btn.grid(row=2, column=0, pady=10)

update_btn = tk.Button(
    window,
    text="Update",
    font=("Poppins", 12, "bold"),
    bg="lightgreen",
    command=update_record
)

update_btn.grid(row=2, column=1)

delete_btn = tk.Button(
    window,
    text="Delete",
    font=("Poppins", 12, "bold"),
    bg="red",
    fg="white",
    command=delete_record
)

delete_btn.grid(row=2, column=2)

clear_btn = tk.Button(
    window,
    text="Clear",
    font=("Poppins", 12, "bold"),
    bg="yellow",
    command=clear_fields
)

clear_btn.grid(row=2, column=3)

table = ttk.Treeview(
    window,
    columns=(
        "Reservation ID",
        "Student Name",
        "Student ID",
        "Room Number",
        "Hours",
        "Rate",
        "Total"
    ),
    show="headings"
)

headings = (
    "Reservation ID",
    "Student Name",
    "Student ID",
    "Room Number",
    "Hours",
    "Rate",
    "Total"
)

for col in headings:
    table.heading(col, text=col)
    table.column(col, width=120)

table.grid(row=3, column=0, columnspan=6, padx=10, pady=10)

table.bind("<ButtonRelease-1>", select_record)

generate_id()
display_records()

window.mainloop()